"""
CADPro Advanced - 개선된 선택 메커니즘
AutoCAD COM API 오류 수정 및 강화된 선택 기능
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import win32com.client
import pythoncom
import math
import json
from datetime import datetime
import os
import time


class CADProAdvanced:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("CADPro Advanced - 개선된 물량 산출")
        self.root.geometry("1200x700")
        
        # AutoCAD 연결
        self.acad = None
        self.doc = None
        self.model = None
        
        # 데이터 저장
        self.all_objects = []
        self.selected_objects = []
        self.filtered_objects = []
        self.calculation_results = {}
        
        # 선택 세트 관리
        self.selection_sets = []
        
        # GUI 구성
        self.setup_gui()
    
    def setup_gui(self):
        """GUI 구성"""
        
        # 노트북 위젯으로 탭 구성
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # 탭 1: 연결 및 선택
        self.tab1 = ttk.Frame(notebook)
        notebook.add(self.tab1, text="연결 및 선택")
        self.setup_connection_tab()
        
        # 탭 2: 계산
        self.tab2 = ttk.Frame(notebook)
        notebook.add(self.tab2, text="물량 계산")
        self.setup_calculation_tab()
        
        # 탭 3: 결과
        self.tab3 = ttk.Frame(notebook)
        notebook.add(self.tab3, text="결과")
        self.setup_result_tab()
    
    def setup_connection_tab(self):
        """연결 및 선택 탭"""
        
        # AutoCAD 연결 프레임
        conn_frame = ttk.LabelFrame(self.tab1, text="AutoCAD 연결", padding="10")
        conn_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(conn_frame, text="AutoCAD 연결", 
                  command=self.connect_autocad, width=20).pack(side=tk.LEFT, padx=5)
        
        self.status_label = ttk.Label(conn_frame, text="연결 대기 중...", foreground="gray")
        self.status_label.pack(side=tk.LEFT, padx=20)
        
        self.drawing_label = ttk.Label(conn_frame, text="도면: -")
        self.drawing_label.pack(side=tk.LEFT, padx=20)
        
        # 선택 방법 프레임
        select_frame = ttk.LabelFrame(self.tab1, text="객체 선택 방법", padding="10")
        select_frame.pack(fill='x', padx=10, pady=5)
        
        # 선택 버튼들
        btn_frame1 = ttk.Frame(select_frame)
        btn_frame1.pack(fill='x', pady=5)
        
        ttk.Button(btn_frame1, text="모든 객체 로드", 
                  command=self.load_all_objects, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame1, text="현재 선택 가져오기", 
                  command=self.get_current_selection, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame1, text="레이어로 선택", 
                  command=self.select_by_layer_dialog, width=20).pack(side=tk.LEFT, padx=5)
        
        btn_frame2 = ttk.Frame(select_frame)
        btn_frame2.pack(fill='x', pady=5)
        
        ttk.Button(btn_frame2, text="타입으로 선택", 
                  command=self.select_by_type_dialog, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame2, text="대화형 선택", 
                  command=self.interactive_selection, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame2, text="선택 초기화", 
                  command=self.clear_selection, width=20).pack(side=tk.LEFT, padx=5)
        
        # 선택 정보 프레임
        info_frame = ttk.LabelFrame(self.tab1, text="선택 정보", padding="10")
        info_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # 정보 표시
        self.info_text = scrolledtext.ScrolledText(info_frame, height=15, width=80)
        self.info_text.pack(fill='both', expand=True)
    
    def setup_calculation_tab(self):
        """계산 탭"""
        
        # 계산 옵션
        option_frame = ttk.LabelFrame(self.tab2, text="계산 옵션", padding="10")
        option_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(option_frame, text="단위:").grid(row=0, column=0, padx=5, pady=5)
        self.unit_var = tk.StringVar(value="mm")
        ttk.Combobox(option_frame, textvariable=self.unit_var, 
                    values=["mm", "m"], width=10).grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(option_frame, text="할증률(%):").grid(row=0, column=2, padx=5, pady=5)
        self.waste_var = tk.StringVar(value="5")
        ttk.Entry(option_frame, textvariable=self.waste_var, width=10).grid(row=0, column=3, padx=5, pady=5)
        
        ttk.Label(option_frame, text="높이(m):").grid(row=0, column=4, padx=5, pady=5)
        self.height_var = tk.StringVar(value="3.0")
        ttk.Entry(option_frame, textvariable=self.height_var, width=10).grid(row=0, column=5, padx=5, pady=5)
        
        # 계산 버튼들
        calc_frame = ttk.LabelFrame(self.tab2, text="계산 실행", padding="10")
        calc_frame.pack(fill='x', padx=10, pady=5)
        
        # 기본 계산
        basic_frame = ttk.Frame(calc_frame)
        basic_frame.pack(fill='x', pady=5)
        
        ttk.Label(basic_frame, text="기본 계산:").pack(side=tk.LEFT, padx=5)
        ttk.Button(basic_frame, text="길이", command=self.calc_length, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(basic_frame, text="면적", command=self.calc_area, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(basic_frame, text="개수", command=self.calc_count, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(basic_frame, text="체적", command=self.calc_volume, width=12).pack(side=tk.LEFT, padx=3)
        
        # 건축 계산
        arch_frame = ttk.Frame(calc_frame)
        arch_frame.pack(fill='x', pady=5)
        
        ttk.Label(arch_frame, text="건축:").pack(side=tk.LEFT, padx=5)
        ttk.Button(arch_frame, text="벽체", command=self.calc_walls, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(arch_frame, text="기둥", command=self.calc_columns, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(arch_frame, text="슬래브", command=self.calc_slabs, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(arch_frame, text="철근", command=self.calc_rebar, width=12).pack(side=tk.LEFT, padx=3)
        
        # 토목 계산
        civil_frame = ttk.Frame(calc_frame)
        civil_frame.pack(fill='x', pady=5)
        
        ttk.Label(civil_frame, text="토목:").pack(side=tk.LEFT, padx=5)
        ttk.Button(civil_frame, text="도로", command=self.calc_road, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(civil_frame, text="배관", command=self.calc_pipes, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(civil_frame, text="토공", command=self.calc_earthwork, width=12).pack(side=tk.LEFT, padx=3)
        
        # 조경 계산
        landscape_frame = ttk.Frame(calc_frame)
        landscape_frame.pack(fill='x', pady=5)
        
        ttk.Label(landscape_frame, text="조경:").pack(side=tk.LEFT, padx=5)
        ttk.Button(landscape_frame, text="데크", command=self.calc_deck, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(landscape_frame, text="수목", command=self.calc_trees, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(landscape_frame, text="포장", command=self.calc_pavement, width=12).pack(side=tk.LEFT, padx=3)
    
    def setup_result_tab(self):
        """결과 탭"""
        
        # 결과 표시
        self.result_text = scrolledtext.ScrolledText(self.tab3, height=25, width=100)
        self.result_text.pack(fill='both', expand=True, padx=10, pady=5)
        
        # 버튼들
        btn_frame = ttk.Frame(self.tab3)
        btn_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(btn_frame, text="결과 지우기", 
                  command=lambda: self.result_text.delete(1.0, tk.END), width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="JSON 저장", 
                  command=self.save_json, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Excel 저장", 
                  command=self.save_excel, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="클립보드 복사", 
                  command=self.copy_to_clipboard, width=15).pack(side=tk.LEFT, padx=5)
    
    # === AutoCAD 연결 및 선택 ===
    
    def connect_autocad(self):
        """AutoCAD 연결"""
        try:
            self.acad = win32com.client.Dispatch("AutoCAD.Application")
            
            if self.acad.Documents.Count == 0:
                messagebox.showwarning("경고", "열린 도면이 없습니다.")
                return
            
            self.doc = self.acad.ActiveDocument
            self.model = self.doc.ModelSpace
            
            # 선택 세트 정리
            self.cleanup_selection_sets()
            
            # UI 업데이트
            self.status_label.config(text="연결됨", foreground="green")
            self.drawing_label.config(text=f"도면: {self.doc.Name}")
            
            # 정보 표시
            info = f"AutoCAD 연결 성공\n"
            info += f"도면: {self.doc.Name}\n"
            info += f"객체 수: {self.model.Count}개\n"
            info += f"레이어 수: {self.doc.Layers.Count}개\n"
            
            self.info_text.delete(1.0, tk.END)
            self.info_text.insert(tk.END, info)
            
            messagebox.showinfo("성공", "AutoCAD 연결 성공!")
            
        except Exception as e:
            messagebox.showerror("오류", f"AutoCAD 연결 실패:\n{str(e)}")
    
    def cleanup_selection_sets(self):
        """기존 선택 세트 정리"""
        try:
            # 기존 선택 세트 삭제
            for i in range(self.doc.SelectionSets.Count - 1, -1, -1):
                try:
                    sel_set = self.doc.SelectionSets.Item(i)
                    if sel_set.Name.startswith("CADPro_"):
                        sel_set.Delete()
                except:
                    pass
        except:
            pass
    
    def load_all_objects(self):
        """모든 객체 로드"""
        if not self.model:
            messagebox.showwarning("경고", "먼저 AutoCAD에 연결하세요.")
            return
        
        try:
            self.all_objects = []
            self.selected_objects = []
            
            # 모든 객체 수집
            for obj in self.model:
                obj_data = self.extract_object_data(obj)
                if obj_data:
                    obj_data['object'] = obj  # COM 객체 참조 저장
                    self.all_objects.append(obj_data)
                    self.selected_objects.append(obj_data)
            
            # 정보 표시
            self.show_selection_info()
            
            messagebox.showinfo("완료", f"{len(self.all_objects)}개 객체를 로드했습니다.")
            
        except Exception as e:
            messagebox.showerror("오류", f"객체 로드 실패:\n{str(e)}")
    
    def get_current_selection(self):
        """현재 AutoCAD에서 선택된 객체 가져오기"""
        if not self.doc:
            messagebox.showwarning("경고", "먼저 AutoCAD에 연결하세요.")
            return
        
        try:
            # PickfirstSelectionSet 사용
            sel_set = self.doc.PickfirstSelectionSet
            
            if sel_set.Count == 0:
                messagebox.showinfo("정보", "선택된 객체가 없습니다.\nAutoCAD에서 객체를 먼저 선택하세요.")
                return
            
            self.selected_objects = []
            
            for obj in sel_set:
                obj_data = self.extract_object_data(obj)
                if obj_data:
                    obj_data['object'] = obj
                    self.selected_objects.append(obj_data)
            
            # 정보 표시
            self.show_selection_info()
            
            messagebox.showinfo("완료", f"{len(self.selected_objects)}개 객체를 가져왔습니다.")
            
        except Exception as e:
            messagebox.showerror("오류", f"선택 가져오기 실패:\n{str(e)}")
    
    def select_by_layer_dialog(self):
        """레이어 선택 대화상자"""
        if not self.doc:
            messagebox.showwarning("경고", "먼저 AutoCAD에 연결하세요.")
            return
        
        # 레이어 목록 가져오기
        layers = []
        for layer in self.doc.Layers:
            layers.append(layer.Name)
        
        # 선택 대화상자
        dialog = tk.Toplevel(self.root)
        dialog.title("레이어 선택")
        dialog.geometry("400x500")
        
        ttk.Label(dialog, text="선택할 레이어:").pack(pady=5)
        
        # 리스트박스
        listbox = tk.Listbox(dialog, selectmode=tk.MULTIPLE, height=20)
        listbox.pack(fill='both', expand=True, padx=10, pady=5)
        
        for layer in sorted(layers):
            listbox.insert(tk.END, layer)
        
        def apply_selection():
            selected_indices = listbox.curselection()
            selected_layers = [listbox.get(i) for i in selected_indices]
            
            if not selected_layers:
                messagebox.showwarning("경고", "레이어를 선택하세요.")
                return
            
            # 선택된 레이어의 객체 수집
            self.selected_objects = []
            
            for obj in self.model:
                if obj.Layer in selected_layers:
                    obj_data = self.extract_object_data(obj)
                    if obj_data:
                        obj_data['object'] = obj
                        self.selected_objects.append(obj_data)
            
            self.show_selection_info()
            dialog.destroy()
            
            messagebox.showinfo("완료", 
                f"{len(selected_layers)}개 레이어에서 {len(self.selected_objects)}개 객체 선택")
        
        ttk.Button(dialog, text="적용", command=apply_selection).pack(pady=10)
    
    def select_by_type_dialog(self):
        """타입 선택 대화상자"""
        if not self.model:
            messagebox.showwarning("경고", "먼저 AutoCAD에 연결하세요.")
            return
        
        # 타입 목록 수집
        types = set()
        for obj in self.model:
            types.add(obj.ObjectName)
        
        # 선택 대화상자
        dialog = tk.Toplevel(self.root)
        dialog.title("타입 선택")
        dialog.geometry("400x500")
        
        ttk.Label(dialog, text="선택할 객체 타입:").pack(pady=5)
        
        # 리스트박스
        listbox = tk.Listbox(dialog, selectmode=tk.MULTIPLE, height=20)
        listbox.pack(fill='both', expand=True, padx=10, pady=5)
        
        for obj_type in sorted(types):
            listbox.insert(tk.END, obj_type)
        
        def apply_selection():
            selected_indices = listbox.curselection()
            selected_types = [listbox.get(i) for i in selected_indices]
            
            if not selected_types:
                messagebox.showwarning("경고", "타입을 선택하세요.")
                return
            
            # 선택된 타입의 객체 수집
            self.selected_objects = []
            
            for obj in self.model:
                if obj.ObjectName in selected_types:
                    obj_data = self.extract_object_data(obj)
                    if obj_data:
                        obj_data['object'] = obj
                        self.selected_objects.append(obj_data)
            
            self.show_selection_info()
            dialog.destroy()
            
            messagebox.showinfo("완료", 
                f"{len(selected_types)}개 타입에서 {len(self.selected_objects)}개 객체 선택")
        
        ttk.Button(dialog, text="적용", command=apply_selection).pack(pady=10)
    
    def interactive_selection(self):
        """대화형 선택 (개선된 방식)"""
        if not self.doc:
            messagebox.showwarning("경고", "먼저 AutoCAD에 연결하세요.")
            return
        
        try:
            # AutoCAD 활성화
            self.acad.Visible = True
            
            # SendCommand 사용하여 SELECT 명령 실행
            self.doc.SendCommand("_SELECT\n")
            
            # 잠시 대기 후 선택된 객체 가져오기
            messagebox.showinfo("선택", 
                "AutoCAD에서 객체를 선택한 후\nEnter를 누르고 이 창에서 OK를 클릭하세요.")
            
            # 선택된 객체 가져오기
            self.get_current_selection()
            
        except Exception as e:
            messagebox.showerror("오류", f"대화형 선택 실패:\n{str(e)}")
    
    def clear_selection(self):
        """선택 초기화"""
        self.selected_objects = []
        self.info_text.delete(1.0, tk.END)
        self.info_text.insert(tk.END, "선택된 객체가 없습니다.")
        messagebox.showinfo("완료", "선택이 초기화되었습니다.")
    
    def extract_object_data(self, obj):
        """객체 데이터 추출"""
        try:
            data = {
                'type': obj.ObjectName,
                'layer': obj.Layer,
                'handle': obj.Handle
            }
            
            # 색상
            try:
                data['color'] = obj.Color
            except:
                pass
            
            # 선 타입
            try:
                data['linetype'] = obj.Linetype
            except:
                pass
            
            # 길이 계산
            if "Line" in obj.ObjectName and "Polyline" not in obj.ObjectName:
                try:
                    start = obj.StartPoint
                    end = obj.EndPoint
                    length = math.sqrt((end[0]-start[0])**2 + (end[1]-start[1])**2 + (end[2]-start[2])**2)
                    data['length'] = length
                    data['start'] = start
                    data['end'] = end
                except:
                    pass
            
            # 폴리라인
            elif "Polyline" in obj.ObjectName:
                try:
                    data['length'] = obj.Length
                    data['closed'] = obj.Closed
                    if obj.Closed and hasattr(obj, 'Area'):
                        data['area'] = obj.Area
                except:
                    pass
            
            # 원
            elif "Circle" in obj.ObjectName:
                try:
                    data['radius'] = obj.Radius
                    data['center'] = obj.Center
                    data['area'] = math.pi * obj.Radius ** 2
                    data['circumference'] = 2 * math.pi * obj.Radius
                except:
                    pass
            
            # 호
            elif "Arc" in obj.ObjectName:
                try:
                    data['radius'] = obj.Radius
                    data['start_angle'] = obj.StartAngle
                    data['end_angle'] = obj.EndAngle
                    data['arc_length'] = obj.ArcLength
                except:
                    pass
            
            # 블록
            elif "BlockReference" in obj.ObjectName:
                try:
                    data['block_name'] = obj.Name
                    data['position'] = obj.InsertionPoint
                    data['scale_x'] = obj.XScaleFactor
                    data['scale_y'] = obj.YScaleFactor
                    data['scale_z'] = obj.ZScaleFactor
                    data['rotation'] = obj.Rotation
                except:
                    pass
            
            # 텍스트
            elif "Text" in obj.ObjectName or "MText" in obj.ObjectName:
                try:
                    data['text'] = obj.TextString
                    data['height'] = obj.Height
                except:
                    pass
            
            # 해치
            elif "Hatch" in obj.ObjectName:
                try:
                    data['pattern'] = obj.PatternName
                    data['area'] = obj.Area
                except:
                    pass
            
            # 치수
            elif "Dimension" in obj.ObjectName:
                try:
                    data['measurement'] = obj.Measurement
                except:
                    pass
            
            return data
            
        except:
            return None
    
    def show_selection_info(self):
        """선택 정보 표시"""
        if not self.selected_objects:
            self.info_text.delete(1.0, tk.END)
            self.info_text.insert(tk.END, "선택된 객체가 없습니다.")
            return
        
        # 타입별 집계
        type_count = {}
        layer_count = {}
        total_length = 0
        total_area = 0
        
        for obj in self.selected_objects:
            # 타입
            obj_type = obj.get('type', 'Unknown')
            type_count[obj_type] = type_count.get(obj_type, 0) + 1
            
            # 레이어
            layer = obj.get('layer', 'Unknown')
            layer_count[layer] = layer_count.get(layer, 0) + 1
            
            # 길이
            if 'length' in obj:
                total_length += obj['length']
            
            # 면적
            if 'area' in obj:
                total_area += obj['area']
        
        # 정보 텍스트 구성
        info = f"=== 선택 정보 ===\n\n"
        info += f"총 선택 객체: {len(self.selected_objects)}개\n\n"
        
        info += "타입별 분포:\n"
        for obj_type, count in sorted(type_count.items(), key=lambda x: x[1], reverse=True):
            info += f"  {obj_type}: {count}개\n"
        
        info += f"\n레이어별 분포 (상위 10개):\n"
        sorted_layers = sorted(layer_count.items(), key=lambda x: x[1], reverse=True)[:10]
        for layer, count in sorted_layers:
            info += f"  {layer}: {count}개\n"
        
        if total_length > 0:
            info += f"\n총 길이: {total_length:.2f}mm ({total_length/1000:.2f}m)\n"
        
        if total_area > 0:
            info += f"총 면적: {total_area:.2f}mm² ({total_area/1000000:.2f}m²)\n"
        
        # 표시
        self.info_text.delete(1.0, tk.END)
        self.info_text.insert(tk.END, info)
    
    # === 계산 메서드 ===
    
    def calc_length(self):
        """길이 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        total = 0
        count = 0
        
        for obj in self.selected_objects:
            if 'length' in obj:
                total += obj['length']
                count += 1
        
        if count == 0:
            result = "길이를 계산할 수 있는 객체가 없습니다."
        else:
            if self.unit_var.get() == "m":
                total = total / 1000
                unit = "m"
            else:
                unit = "mm"
            
            result = f"=== 길이 계산 ===\n\n"
            result += f"객체 수: {count}개\n"
            result += f"총 길이: {total:.2f} {unit}\n"
        
        self.display_result(result)
    
    def calc_area(self):
        """면적 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        total = 0
        count = 0
        
        for obj in self.selected_objects:
            if 'area' in obj:
                total += obj['area']
                count += 1
        
        if count == 0:
            result = "면적을 계산할 수 있는 객체가 없습니다."
        else:
            if self.unit_var.get() == "m":
                total = total / 1000000
                unit = "m²"
            else:
                unit = "mm²"
            
            result = f"=== 면적 계산 ===\n\n"
            result += f"객체 수: {count}개\n"
            result += f"총 면적: {total:.2f} {unit}\n"
            
            if self.unit_var.get() == "m":
                result += f"평 환산: {total/3.3:.2f} 평\n"
        
        self.display_result(result)
    
    def calc_count(self):
        """개수 집계"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        # 타입별 집계
        type_count = {}
        for obj in self.selected_objects:
            obj_type = obj.get('type', 'Unknown')
            type_count[obj_type] = type_count.get(obj_type, 0) + 1
        
        result = f"=== 개수 집계 ===\n\n"
        result += f"총 객체 수: {len(self.selected_objects)}개\n\n"
        result += "타입별 개수:\n"
        
        for obj_type, count in sorted(type_count.items(), key=lambda x: x[1], reverse=True):
            result += f"  {obj_type}: {count}개\n"
        
        self.display_result(result)
    
    def calc_volume(self):
        """체적 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        # 면적 * 높이로 체적 추정
        total_area = 0
        count = 0
        
        for obj in self.selected_objects:
            if 'area' in obj:
                total_area += obj['area']
                count += 1
        
        if count == 0:
            result = "체적을 계산할 수 있는 면적 객체가 없습니다."
        else:
            height = float(self.height_var.get())
            
            if self.unit_var.get() == "m":
                volume = (total_area / 1000000) * height
                unit = "m³"
            else:
                volume = total_area * height
                unit = "mm³"
            
            result = f"=== 체적 계산 ===\n\n"
            result += f"면적 객체: {count}개\n"
            result += f"높이/두께: {height}m\n"
            result += f"총 체적: {volume:.3f} {unit}\n"
        
        self.display_result(result)
    
    def calc_walls(self):
        """벽체 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        total_length = 0
        count = 0
        
        for obj in self.selected_objects:
            if 'length' in obj:
                total_length += obj['length']
                count += 1
        
        if count == 0:
            result = "벽체로 계산할 선 객체가 없습니다."
        else:
            length_m = total_length / 1000
            height = float(self.height_var.get())
            thickness = 0.2  # 200mm
            
            wall_area = length_m * height
            wall_volume = wall_area * thickness
            concrete = wall_volume * 1.05  # 5% 할증
            
            result = f"=== 벽체 계산 ===\n\n"
            result += f"벽체 길이: {length_m:.2f}m\n"
            result += f"벽체 높이: {height}m\n"
            result += f"벽체 두께: {thickness}m\n\n"
            result += f"벽체 면적: {wall_area:.2f}m²\n"
            result += f"벽체 체적: {wall_volume:.3f}m³\n"
            result += f"콘크리트: {concrete:.3f}m³ (5% 할증)\n"
        
        self.display_result(result)
        self.calculation_results['벽체'] = {
            '길이': length_m,
            '면적': wall_area,
            '체적': wall_volume,
            '콘크리트': concrete
        }
    
    def calc_columns(self):
        """기둥 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        circles = 0
        blocks = 0
        
        for obj in self.selected_objects:
            if "Circle" in obj.get('type', ''):
                circles += 1
            elif "Block" in obj.get('type', ''):
                blocks += 1
        
        total = circles + blocks
        
        if total == 0:
            result = "기둥으로 계산할 객체(원, 블록)가 없습니다."
        else:
            foundation = total * 0.096  # 400x400x600
            concrete = foundation * 1.05
            
            result = f"=== 기둥 계산 ===\n\n"
            result += f"원형 객체: {circles}개\n"
            result += f"블록 객체: {blocks}개\n"
            result += f"총 기둥: {total}개\n\n"
            result += f"기초 체적: {foundation:.3f}m³\n"
            result += f"콘크리트: {concrete:.3f}m³ (5% 할증)\n"
        
        self.display_result(result)
        self.calculation_results['기둥'] = {
            '개수': total,
            '기초': foundation,
            '콘크리트': concrete
        }
    
    def calc_slabs(self):
        """슬래브 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        total_area = 0
        count = 0
        
        for obj in self.selected_objects:
            if 'area' in obj and obj.get('closed', False):
                total_area += obj['area']
                count += 1
        
        if count == 0:
            result = "슬래브로 계산할 닫힌 도형이 없습니다."
        else:
            area_m2 = total_area / 1000000
            thickness = 0.15  # 150mm
            volume = area_m2 * thickness
            concrete = volume * 1.05
            rebar = area_m2 * 0.015  # ton/m²
            
            result = f"=== 슬래브 계산 ===\n\n"
            result += f"슬래브 면적: {area_m2:.2f}m²\n"
            result += f"슬래브 두께: {thickness}m\n\n"
            result += f"콘크리트: {concrete:.3f}m³\n"
            result += f"철근: {rebar:.3f}ton\n"
        
        self.display_result(result)
    
    def calc_rebar(self):
        """철근 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        total_length = 0
        count = 0
        
        for obj in self.selected_objects:
            if 'length' in obj:
                total_length += obj['length']
                count += 1
        
        if count == 0:
            result = "철근으로 계산할 선 객체가 없습니다."
        else:
            length_m = total_length / 1000
            
            # D16 기준
            weight = length_m * 1.56  # kg
            weight_ton = weight / 1000
            
            result = f"=== 철근 계산 ===\n\n"
            result += f"철근 개수: {count}개\n"
            result += f"총 길이: {length_m:.2f}m\n\n"
            result += f"중량 (D16 기준): {weight:.2f}kg\n"
            result += f"중량: {weight_ton:.3f}ton\n"
            result += f"할증 3%: {weight_ton*1.03:.3f}ton\n"
        
        self.display_result(result)
    
    def calc_road(self):
        """도로 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        total_area = 0
        count = 0
        
        for obj in self.selected_objects:
            if 'area' in obj:
                total_area += obj['area']
                count += 1
        
        if count == 0:
            result = "도로로 계산할 면적 객체가 없습니다."
        else:
            area_m2 = total_area / 1000000
            asphalt = area_m2 * 0.1  # 100mm
            base = area_m2 * 0.15  # 150mm
            
            result = f"=== 도로 계산 ===\n\n"
            result += f"도로 면적: {area_m2:.2f}m²\n\n"
            result += f"아스팔트 (100mm): {asphalt:.3f}m³\n"
            result += f"기층 (150mm): {base:.3f}m³\n"
        
        self.display_result(result)
    
    def calc_pipes(self):
        """배관 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        total_length = 0
        count = 0
        
        for obj in self.selected_objects:
            if 'length' in obj:
                total_length += obj['length']
                count += 1
        
        if count == 0:
            result = "배관으로 계산할 선 객체가 없습니다."
        else:
            length_m = total_length / 1000
            joints = int(length_m / 6)  # 6m당 1개
            
            result = f"=== 배관 계산 ===\n\n"
            result += f"배관 개수: {count}개\n"
            result += f"총 연장: {length_m:.2f}m\n\n"
            result += f"이음관: {joints}개 (6m당 1개)\n"
        
        self.display_result(result)
    
    def calc_earthwork(self):
        """토공 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        total_area = 0
        count = 0
        
        for obj in self.selected_objects:
            if 'area' in obj:
                total_area += obj['area']
                count += 1
        
        if count == 0:
            result = "토공으로 계산할 면적 객체가 없습니다."
        else:
            area_m2 = total_area / 1000000
            depth = 1.0  # 1m 깊이 가정
            volume = area_m2 * depth
            
            result = f"=== 토공 계산 ===\n\n"
            result += f"작업 면적: {area_m2:.2f}m²\n"
            result += f"평균 깊이: {depth}m\n\n"
            result += f"토공량: {volume:.3f}m³\n"
            result += f"할증 20%: {volume*1.2:.3f}m³\n"
        
        self.display_result(result)
    
    def calc_deck(self):
        """데크 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        # 장선 (선)
        joist_length = 0
        joist_count = 0
        
        # 데크 면적
        deck_area = 0
        area_count = 0
        
        for obj in self.selected_objects:
            if 'length' in obj:
                joist_length += obj['length']
                joist_count += 1
            if 'area' in obj:
                deck_area += obj['area']
                area_count += 1
        
        result = f"=== 데크 계산 ===\n\n"
        
        if joist_count > 0:
            length_m = joist_length / 1000
            volume = length_m * 0.038 * 0.235  # 2x10
            
            result += "[장선]\n"
            result += f"  개수: {joist_count}개\n"
            result += f"  총 길이: {length_m:.2f}m\n"
            result += f"  목재 체적: {volume:.3f}m³ (2x10)\n\n"
        
        if area_count > 0:
            area_m2 = deck_area / 1000000
            waste = float(self.waste_var.get()) / 100
            board_area = area_m2 * (1 + waste)
            screws = int(area_m2 * 8)
            
            result += "[데크보드]\n"
            result += f"  면적: {area_m2:.2f}m²\n"
            result += f"  필요 보드: {board_area:.2f}m² (할증 포함)\n"
            result += f"  데크 스크류: {screws}개\n"
        
        if joist_count == 0 and area_count == 0:
            result += "데크 계산할 객체가 없습니다."
        
        self.display_result(result)
    
    def calc_trees(self):
        """수목 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        circles = 0
        blocks = 0
        
        for obj in self.selected_objects:
            if "Circle" in obj.get('type', ''):
                circles += 1
            elif "Block" in obj.get('type', ''):
                blocks += 1
        
        total = circles + blocks
        
        if total == 0:
            result = "수목으로 계산할 객체(원, 블록)가 없습니다."
        else:
            result = f"=== 수목 계산 ===\n\n"
            result += f"원형 객체: {circles}개\n"
            result += f"블록 객체: {blocks}개\n"
            result += f"총 수목: {total}주\n"
        
        self.display_result(result)
    
    def calc_pavement(self):
        """포장 계산"""
        if not self.selected_objects:
            messagebox.showwarning("경고", "먼저 객체를 선택하세요.")
            return
        
        total_area = 0
        count = 0
        
        for obj in self.selected_objects:
            if 'area' in obj:
                total_area += obj['area']
                count += 1
        
        if count == 0:
            result = "포장으로 계산할 면적 객체가 없습니다."
        else:
            area_m2 = total_area / 1000000
            blocks = area_m2 * 50  # 50개/m²
            sand = area_m2 * 0.03  # 30mm
            
            result = f"=== 포장 계산 ===\n\n"
            result += f"포장 면적: {area_m2:.2f}m²\n\n"
            result += f"보도블록: {int(blocks)}개 (50개/m²)\n"
            result += f"모래: {sand:.3f}m³\n"
        
        self.display_result(result)
    
    # === 유틸리티 ===
    
    def display_result(self, text):
        """결과 표시"""
        self.result_text.insert(tk.END, "\n" + text + "\n")
        self.result_text.insert(tk.END, f"계산 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        self.result_text.insert(tk.END, "-" * 60 + "\n")
        self.result_text.see(tk.END)
    
    def save_json(self):
        """JSON 저장"""
        if not self.calculation_results:
            messagebox.showwarning("경고", "저장할 결과가 없습니다.")
            return
        
        filename = f"cadpro_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        try:
            output = {
                "도면": self.doc.Name if self.doc else "Unknown",
                "일시": datetime.now().isoformat(),
                "선택객체수": len(self.selected_objects),
                "계산결과": self.calculation_results
            }
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(output, f, ensure_ascii=False, indent=2)
            
            messagebox.showinfo("성공", f"저장 완료: {filename}")
            
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {str(e)}")
    
    def save_excel(self):
        """Excel 저장 (openpyxl 필요)"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "물량산출"
            
            # 헤더
            ws['A1'] = "CADPro 물량산출 결과"
            ws['A2'] = f"도면: {self.doc.Name if self.doc else 'Unknown'}"
            ws['A3'] = f"일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # 결과
            row = 5
            for category, data in self.calculation_results.items():
                ws[f'A{row}'] = category
                row += 1
                for key, value in data.items():
                    ws[f'B{row}'] = key
                    ws[f'C{row}'] = value
                    row += 1
                row += 1
            
            filename = f"cadpro_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            messagebox.showinfo("성공", f"Excel 저장 완료: {filename}")
            
        except ImportError:
            messagebox.showerror("오류", "openpyxl이 설치되지 않았습니다.\npip install openpyxl")
        except Exception as e:
            messagebox.showerror("오류", f"Excel 저장 실패: {str(e)}")
    
    def copy_to_clipboard(self):
        """클립보드 복사"""
        try:
            text = self.result_text.get(1.0, tk.END)
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            messagebox.showinfo("성공", "클립보드에 복사되었습니다.")
        except:
            pass
    
    def on_closing(self):
        """창 닫기 이벤트 처리"""
        try:
            # 선택 세트 정리
            if self.doc:
                self.cleanup_selection_sets()
            
            # 창 파괴
            self.root.quit()
            self.root.destroy()
        except:
            self.root.destroy()
    
    def run(self):
        """프로그램 실행"""
        # 종료 이벤트 핸들러 등록
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        try:
            self.root.mainloop()
        except KeyboardInterrupt:
            self.on_closing()


def main():
    try:
        app = CADProAdvanced()
        app.run()
    except Exception as e:
        print(f"프로그램 오류: {e}")
        import sys
        sys.exit(1)


if __name__ == "__main__":
    main()